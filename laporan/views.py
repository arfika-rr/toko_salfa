from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
from kasir.models import Transaksi, DetailTransaksi
from produk.models import Produk
from piutang.models import Piutang
import json
import csv

def get_data_laporan(bulan, tahun):
    transaksi = Transaksi.objects.filter(
        tanggal__month=bulan,
        tanggal__year=tahun,
    )
    total_omzet = transaksi.aggregate(s=Sum('total'))['s'] or 0
    total_transaksi = transaksi.count()

    produk_terlaris = DetailTransaksi.objects.filter(
        transaksi__tanggal__month=bulan,
        transaksi__tanggal__year=tahun,
    ).select_related('produk').values(
        'produk__nama', 'produk__harga_beli'
    ).annotate(
        total_terjual=Sum('jumlah'),
        total_omzet=Sum('subtotal'),
    ).order_by('-total_terjual')

    data_produk = []
    total_profit = 0
    for p in produk_terlaris:
        if p['produk__nama'] is None or p['produk__harga_beli'] is None:
            continue
        profit = (p['total_omzet'] - (p['produk__harga_beli'] * p['total_terjual']))
        total_profit += profit
        data_produk.append({
            'nama': p['produk__nama'],
            'total_terjual': p['total_terjual'],
            'total_omzet': float(p['total_omzet']),
            'profit': float(profit),
        }) 

    piutang_belum = Piutang.objects.exclude(status='lunas')
    total_piutang = piutang_belum.aggregate(s=Sum('nominal'))['s'] or 0

    return {
        'total_omzet': total_omzet,
        'total_transaksi': total_transaksi,
        'total_profit': total_profit,
        'total_piutang': total_piutang,
        'data_produk': data_produk,
    }


@login_required
def dashboard_laporan(request):
    hari_ini = timezone.now().date()
    bulan = int(request.GET.get('bulan', hari_ini.month))
    tahun = int(request.GET.get('tahun', hari_ini.year))

    data = get_data_laporan(bulan, tahun)

    # Grafik omzet harian
    omzet_harian = Transaksi.objects.filter(
        tanggal__month=bulan,
        tanggal__year=tahun,
    ).annotate(tgl=TruncDate('tanggal')).values('tgl').annotate(
        omzet=Sum('total')
    ).order_by('tgl')

    bulan_indo = {
    1:'Jan', 2:'Feb', 3:'Mar', 4:'Apr', 5:'Mei', 6:'Jun',
    7:'Jul', 8:'Agu', 9:'Sep', 10:'Okt', 11:'Nov', 12:'Des'
    }
    grafik_label = [f"{o['tgl'].day} {bulan_indo[o['tgl'].month]}" for o in omzet_harian]
    grafik_data = [float(o['omzet']) for o in omzet_harian]

    bulan_list = [
        (1,'Januari'),(2,'Februari'),(3,'Maret'),(4,'April'),
        (5,'Mei'),(6,'Juni'),(7,'Juli'),(8,'Agustus'),
        (9,'September'),(10,'Oktober'),(11,'November'),(12,'Desember'),
    ]
    tahun_list = list(range(hari_ini.year, hari_ini.year - 3, -1))
    nama_bulan = dict(bulan_list)[bulan]

    return render(request, 'laporan/dashboard.html', {
        **data,
        'grafik_label': json.dumps(grafik_label),
        'grafik_data': json.dumps(grafik_data),
        'bulan_list': bulan_list,
        'tahun_list': tahun_list,
        'bulan_aktif': bulan,
        'tahun_aktif': tahun,
        'nama_bulan': nama_bulan,
    })


@login_required
def export_laporan_csv(request):
    hari_ini = timezone.now().date()
    bulan = int(request.GET.get('bulan', hari_ini.month))
    tahun = int(request.GET.get('tahun', hari_ini.year))
    data = get_data_laporan(bulan, tahun)

    bulan_list = {
        1:'Januari',2:'Februari',3:'Maret',4:'April',
        5:'Mei',6:'Juni',7:'Juli',8:'Agustus',
        9:'September',10:'Oktober',11:'November',12:'Desember'
    }

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="laporan_{bulan}_{tahun}.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Laporan Penjualan - Toko Salfa'])
    writer.writerow([f"Periode: {bulan_list[bulan]} {tahun}"])
    writer.writerow([])
    writer.writerow(['Total Transaksi', data['total_transaksi']])
    writer.writerow(['Total Omzet', f"{float(data['total_omzet']):,.0f}".replace(',', '.')])
    writer.writerow(['Total Profit', f"{float(data['total_profit']):,.0f}".replace(',', '.')])
    writer.writerow(['Total Piutang Belum Lunas', f"{float(data['total_piutang']):,.0f}".replace(',', '.')])
    writer.writerow([])
    writer.writerow(['No', 'Nama Produk', 'Jumlah Terjual', 'Total Omzet (Rp)', 'Total Profit (Rp)'])

    for i, p in enumerate(data['data_produk'], 1):
        writer.writerow([i, p['nama'], p['total_terjual'], f"{p['total_omzet']:,.0f}", f"{p['profit']:,.0f}"])

    return response


@login_required
def export_laporan_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hari_ini = timezone.now().date()
    bulan = int(request.GET.get('bulan', hari_ini.month))
    tahun = int(request.GET.get('tahun', hari_ini.year))
    data = get_data_laporan(bulan, tahun)

    bulan_list = {
        1:'Januari',2:'Februari',3:'Maret',4:'April',
        5:'Mei',6:'Juni',7:'Juli',8:'Agustus',
        9:'September',10:'Oktober',11:'November',12:'Desember'
    }
    nama_bulan = bulan_list[bulan]

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    # Judul
    ws.merge_cells('A1:E1')
    ws['A1'] = 'LAPORAN PENJUALAN - TOKO SALFA'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Periode: {nama_bulan} {tahun}"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Ringkasan
    ws['A4'] = 'Total Transaksi'
    ws['B4'] = data['total_transaksi']
    ws['A5'] = 'Total Omzet'
    ws['B5'] = float(data['total_omzet'])
    ws['A6'] = 'Total Profit'
    ws['B6'] = float(data['total_profit'])
    ws['A7'] = 'Total Piutang Belum Lunas'
    ws['B7'] = float(data['total_piutang'])
    for row in range(4, 8):
        ws.cell(row=row, column=1).font = Font(bold=True)

    # Header tabel produk
    headers = ['No', 'Nama Produk', 'Jumlah Terjual', 'Total Omzet (Rp)', 'Total Profit (Rp)']
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, p in enumerate(data['data_produk'], 10):
        ws.cell(row=row, column=1, value=row-9)
        ws.cell(row=row, column=2, value=p['nama'])
        ws.cell(row=row, column=3, value=p['total_terjual'])
        ws.cell(row=row, column=4, value=p['total_omzet'])
        ws.cell(row=row, column=5, value=p['profit'])
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="laporan_{bulan}_{tahun}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_laporan_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    hari_ini = timezone.now().date()
    bulan = int(request.GET.get('bulan', hari_ini.month))
    tahun = int(request.GET.get('tahun', hari_ini.year))
    data = get_data_laporan(bulan, tahun)

    bulan_list = {
        1:'Januari',2:'Februari',3:'Maret',4:'April',
        5:'Mei',6:'Juni',7:'Juli',8:'Agustus',
        9:'September',10:'Oktober',11:'November',12:'Desember'
    }
    nama_bulan = bulan_list[bulan]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, spaceAfter=8, textColor=colors.grey)
    elements.append(Paragraph('LAPORAN PENJUALAN', title_style))
    elements.append(Paragraph(f"Toko Salfa — {nama_bulan} {tahun}", sub_style))

    # Ringkasan
    ringkasan = [
        ['Total Transaksi', str(data['total_transaksi'])],
        ['Total Omzet', f"Rp{float(data['total_omzet']):,.0f}".replace(',', '.')],
        ['Total Profit', f"Rp{float(data['total_profit']):,.0f}".replace(',', '.')],
        ['Total Piutang Belum Lunas', f"Rp{float(data['total_piutang']):,.0f}".replace(',', '.')],
    ]
    t_ringkasan = Table(ringkasan, colWidths=[6*cm, 5*cm])
    t_ringkasan.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.HexColor('#EBF5FB'), colors.white]),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(t_ringkasan)
    elements.append(Spacer(1, 0.5*cm))

    # Tabel produk terlaris
    elements.append(Paragraph('Produk Terlaris', styles['Heading2']))
    table_data = [['No', 'Nama Produk', 'Jumlah Terjual', 'Total Omzet (Rp)', 'Total Profit (Rp)']]
    for i, p in enumerate(data['data_produk'], 1):
        table_data.append([
            str(i), p['nama'], str(p['total_terjual']),
            f"{p['total_omzet']:,.0f}", f"{p['profit']:,.0f}",
        ])

    table = Table(table_data, colWidths=[1*cm, 7*cm, 3.5*cm, 4*cm, 4*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A5276')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="laporan_{bulan}_{tahun}.pdf"'
    return response