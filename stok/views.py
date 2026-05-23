from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from .models import StokMasuk
from produk.models import Produk
import csv

@login_required
def list_stok(request):
    histori = StokMasuk.objects.select_related('produk', 'dicatat_oleh').order_by('-tanggal')
    produk_semua = Produk.objects.select_related('kategori').filter(aktif=True).order_by('nama')
    return render(request, 'stok/list.html', {
        'histori': histori,
        'produk_semua': produk_semua,
    })

@login_required
def tambah_stok(request):
    produk = Produk.objects.filter(aktif=True).order_by('nama')
    if request.method == 'POST':
        p = Produk.objects.get(pk=request.POST['produk'])
        jumlah = int(request.POST['jumlah'])
        StokMasuk.objects.create(
            produk=p,
            jumlah=jumlah,
            harga_beli=request.POST['harga_beli'],
            dicatat_oleh=request.user,
            keterangan=request.POST.get('keterangan', ''),
        )
        p.stok += jumlah
        p.harga_beli = request.POST['harga_beli']
        p.save()
        messages.success(request, f'Stok {p.nama} berhasil ditambah {jumlah} {p.satuan}.')
        return redirect('stok:list')
    return render(request, 'stok/form.html', {'produk': produk})


def get_data_stok():
    """Helper: ambil data stok semua produk"""
    produk = Produk.objects.select_related('kategori').filter(aktif=True).order_by('kategori__nama', 'nama')
    data = []
    grand_total_kuantitas = 0
    grand_total_modal = 0
    for p in produk:
        total_modal = p.harga_beli * p.stok
        grand_total_kuantitas += p.stok
        grand_total_modal += total_modal
        data.append({
            'nama': p.nama,
            'kategori': p.kategori.nama if p.kategori else '-',
            'harga_beli': float(p.harga_beli),
            'harga_jual': float(p.harga_jual),
            'stok': p.stok,
            'satuan': p.satuan,
            'total_modal': float(total_modal),
        })
    return data, grand_total_kuantitas, float(grand_total_modal)


@login_required
def export_stok_csv(request):
    data, grand_total_kuantitas, grand_total_modal = get_data_stok()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="laporan_stok.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Laporan Stok Barang - Toko Salfa'])
    writer.writerow([])
    writer.writerow(['No', 'Nama Barang', 'Kategori', 'Harga Beli (Rp)', 'Harga Jual (Rp)', 'Stok', 'Satuan', 'Total Modal (Rp)'])

    for i, d in enumerate(data, 1):
        writer.writerow([
            i, d['nama'], d['kategori'],
            f"{d['harga_beli']:,.0f}", f"{d['harga_jual']:,.0f}",
            d['stok'], d['satuan'],
            f"{d['total_modal']:,.0f}",
        ])

    writer.writerow([])
    writer.writerow(['', '', '', '', '', 'Grand Total Kuantitas', grand_total_kuantitas, ''])
    writer.writerow(['', '', '', '', '', 'Grand Total Modal', '', f"{grand_total_modal:,.0f}"])

    return response


@login_required
def export_stok_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    data, grand_total_kuantitas, grand_total_modal = get_data_stok()

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Stok"

    ws.merge_cells('A1:H1')
    ws['A1'] = 'LAPORAN STOK BARANG - TOKO SALFA'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Nama Barang', 'Kategori', 'Harga Beli (Rp)', 'Harga Jual (Rp)', 'Stok', 'Satuan', 'Total Modal (Rp)']
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, d in enumerate(data, 4):
        ws.cell(row=row, column=1, value=row-3)
        ws.cell(row=row, column=2, value=d['nama'])
        ws.cell(row=row, column=3, value=d['kategori'])
        ws.cell(row=row, column=4, value=d['harga_beli'])
        ws.cell(row=row, column=5, value=d['harga_jual'])
        ws.cell(row=row, column=6, value=d['stok'])
        ws.cell(row=row, column=7, value=d['satuan'])
        ws.cell(row=row, column=8, value=d['total_modal'])
        if row % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')

    last_row = len(data) + 4
    ws.cell(row=last_row+1, column=5, value='Grand Total Kuantitas').font = Font(bold=True)
    ws.cell(row=last_row+1, column=6, value=grand_total_kuantitas).font = Font(bold=True)
    ws.cell(row=last_row+2, column=5, value='Grand Total Modal').font = Font(bold=True)
    ws.cell(row=last_row+2, column=8, value=grand_total_modal).font = Font(bold=True)

    col_widths = [5, 30, 18, 18, 18, 10, 10, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="laporan_stok.xlsx"'
    wb.save(response)
    return response


@login_required
def export_stok_pdf(request):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    data, grand_total_kuantitas, grand_total_modal = get_data_stok()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, spaceAfter=12, textColor=colors.grey)
    elements.append(Paragraph('LAPORAN STOK BARANG', title_style))
    elements.append(Paragraph('Toko Salfa', sub_style))

    table_data = [['No', 'Nama Barang', 'Kategori', 'Harga Beli (Rp)', 'Harga Jual (Rp)', 'Stok', 'Satuan', 'Total Modal (Rp)']]
    for i, d in enumerate(data, 1):
        table_data.append([
            str(i), d['nama'], d['kategori'],
            f"{d['harga_beli']:,.0f}", f"{d['harga_jual']:,.0f}",
            str(d['stok']), d['satuan'],
            f"{d['total_modal']:,.0f}",
        ])

    table_data.append(['', '', '', '', '', '', '', ''])
    table_data.append(['', '', '', '', 'Grand Total Kuantitas', str(grand_total_kuantitas), '', ''])
    table_data.append(['', '', '', '', 'Grand Total Modal', '', '', f"{grand_total_modal:,.0f}"])

    table = Table(table_data,
        colWidths=[1*cm, 5.5*cm, 3*cm, 3.5*cm, 3.5*cm, 2*cm, 2*cm, 3.5*cm],
        repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A5276')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-4), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID', (0,0), (-1,-4), 0.5, colors.grey),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor('#D6EAF8')),
        ('PADDING', (0,0), (-1,-1), 5),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laporan_stok.pdf"'
    return response