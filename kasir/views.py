from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from .models import Transaksi, DetailTransaksi
from produk.models import Produk
from piutang.models import Pelanggan, Piutang
import json
import csv

def fmt_rp(value):
    return f"Rp {float(value):,.0f}".replace(',', '.')

def get_transaksi_filter(request):
    dari = request.GET.get('dari')
    sampai = request.GET.get('sampai')
    transaksi = Transaksi.objects.select_related('kasir').order_by('-tanggal')
    if dari:
        transaksi = transaksi.filter(tanggal__date__gte=dari)
    if sampai:
        transaksi = transaksi.filter(tanggal__date__lte=sampai)
    return transaksi, dari, sampai


def get_ringkasan(transaksi):
    from collections import defaultdict
    ringkasan_produk = defaultdict(lambda: {'jumlah': 0, 'total': 0, 'profit': 0})
    grand_total = 0
    total_profit = 0

    for t in transaksi:
        detail = t.detail.select_related('produk').all()
        grand_total += t.total
        for d in detail:
            if d.produk is None:
                continue
            profit_item = (d.harga_satuan - d.produk.harga_beli) * d.jumlah
            total_profit += profit_item
            nama = d.produk.nama
            ringkasan_produk[nama]['jumlah'] += d.jumlah
            ringkasan_produk[nama]['total'] += float(d.subtotal)
            ringkasan_produk[nama]['profit'] += float(profit_item)

    return ringkasan_produk, grand_total, total_profit


@login_required
def kasir(request):
    produk = Produk.objects.filter(aktif=True).select_related('kategori').order_by('nama')
    pelanggan = Pelanggan.objects.all().order_by('nama')
    return render(request, 'kasir/kasir.html', {
        'produk': produk,
        'pelanggan': pelanggan,
    })


@login_required
def cari_barcode(request):
    barcode = request.GET.get('barcode', '')
    try:
        produk = Produk.objects.get(barcode=barcode, aktif=True)
        return JsonResponse({
            'found': True,
            'id': produk.id,
            'nama': produk.nama,
            'harga': float(produk.harga_jual),
            'stok': produk.stok,
            'satuan': produk.satuan,
        })
    except Produk.DoesNotExist:
        return JsonResponse({'found': False})


@login_required
def proses_transaksi(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        items = data.get('items', [])
        metode = data.get('metode', 'tunai')
        bayar = int(data.get('bayar', 0))
        pelanggan_id = data.get('pelanggan_id')

        if not items:
            return JsonResponse({'success': False, 'error': 'Keranjang kosong'})

        now = timezone.now()
        kode = f"TRX{now.strftime('%Y%m%d%H%M%S')}"
        total = sum(item['harga'] * item['jumlah'] for item in items)
        kembalian = bayar - total if metode == 'tunai' else 0

        transaksi = Transaksi.objects.create(
            kasir=request.user,
            kode_transaksi=kode,
            total=total,
            bayar=bayar,
            kembalian=kembalian,
            metode_bayar=metode,
        )

        for item in items:
            produk = Produk.objects.get(pk=item['id'])
            DetailTransaksi.objects.create(
                transaksi=transaksi,
                produk=produk,
                jumlah=item['jumlah'],
                harga_satuan=item['harga'],
            )
            produk.stok -= item['jumlah']
            produk.save()

        if metode == 'piutang' and pelanggan_id:
            pelanggan = Pelanggan.objects.get(pk=pelanggan_id)
            Piutang.objects.create(
                pelanggan=pelanggan,
                transaksi=transaksi,
                nominal=total,
                catatan=f"Dari transaksi {kode}",
            )

        return JsonResponse({
            'success': True,
            'kode': kode,
            'total': total,
            'kembalian': kembalian,
        })

    return JsonResponse({'success': False})


@login_required
def riwayat_transaksi(request):
    transaksi, dari, sampai = get_transaksi_filter(request)

    data_transaksi = []
    for t in transaksi:
        detail = t.detail.select_related('produk').all()
        keuntungan = sum(
            (d.harga_satuan - d.produk.harga_beli) * d.jumlah
            for d in detail
            if d.produk is not None
        )
        data_transaksi.append({'obj': t, 'keuntungan': keuntungan})

    total_omzet = sum(d['obj'].total for d in data_transaksi)
    total_keuntungan = sum(d['keuntungan'] for d in data_transaksi)

    return render(request, 'kasir/riwayat.html', {
        'data_transaksi': data_transaksi,
        'dari': dari,
        'sampai': sampai,
        'total_omzet': total_omzet,
        'total_keuntungan': total_keuntungan,
    })


@login_required
def detail_transaksi(request, pk):
    transaksi = get_object_or_404(Transaksi, pk=pk)
    detail = transaksi.detail.select_related('produk').all()
    return render(request, 'kasir/detail.html', {
        'transaksi': transaksi,
        'detail': detail,
    })


@login_required
def batal_transaksi(request, pk):
    transaksi = get_object_or_404(Transaksi, pk=pk)

    # Cek apakah sudah dibatalkan
    if transaksi.status == 'batal':
        messages.error(request, 'Transaksi ini sudah dibatalkan sebelumnya.')
        return redirect('kasir:riwayat')

    if request.method == 'POST':
        alasan = request.POST.get('alasan', '').strip()
        if not alasan:
            messages.error(request, 'Alasan pembatalan harus diisi.')
            return redirect('kasir:detail', pk=pk)

        # Kembalikan stok
        for d in transaksi.detail.select_related('produk').all():
            produk = d.produk
            produk.stok += d.jumlah
            produk.save()

        # Batalkan piutang jika ada
        if transaksi.metode_bayar == 'piutang':
            try:
                transaksi.piutang.delete()
            except:
                pass

        # Update status transaksi
        transaksi.status = 'batal'
        transaksi.alasan_batal = alasan
        transaksi.save()

        messages.success(request, f'Transaksi {transaksi.kode_transaksi} berhasil dibatalkan.')
        return redirect('kasir:riwayat')

    return render(request, 'kasir/batal.html', {'transaksi': transaksi})

@login_required
def hapus_transaksi(request, pk):
    transaksi = get_object_or_404(Transaksi, pk=pk)
    if request.method == 'POST':
        kode = transaksi.kode_transaksi
        # Kembalikan stok jika transaksi belum dibatalkan
        if transaksi.status == 'selesai':
            for d in transaksi.detail.select_related('produk').all():
                if d.produk:
                    d.produk.stok += d.jumlah
                    d.produk.save()
        # Hapus piutang terkait jika ada
        try:
            transaksi.piutang.delete()
        except:
            pass
        transaksi.delete()
        messages.success(request, f'Transaksi {kode} berhasil dihapus.')
    return redirect('kasir:riwayat')

@login_required
def export_csv(request):
    transaksi, dari, sampai = get_transaksi_filter(request)
    ringkasan, grand_total, total_profit = get_ringkasan(transaksi)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="laporan_penjualan.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['Laporan Penjualan - Toko Salfa'])
    writer.writerow([f'Periode: {dari or "Semua"} s/d {sampai or "Semua"}'])
    writer.writerow([])
    writer.writerow(['No', 'Nama Produk', 'Jumlah Terjual', 'Total Penjualan (Rp)', 'Total Profit (Rp)'])

    for i, (nama, val) in enumerate(ringkasan.items(), 1):
        writer.writerow([i, nama, val['jumlah'], f"{val['total']:,.0f}", f"{val['profit']:,.0f}"])

    writer.writerow([])
    writer.writerow(['', '', '', 'Grand Total', f"{float(grand_total):,.0f}"]).replace(',', '.')
    writer.writerow(['', '', '', 'Total Profit', f"{float(total_profit):,.0f}"]).replace(',', '.')

    return response


@login_required
def export_excel(request):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    transaksi, dari, sampai = get_transaksi_filter(request)
    ringkasan, grand_total, total_profit = get_ringkasan(transaksi)

    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Penjualan"

    ws.merge_cells('A1:E1')
    ws['A1'] = 'LAPORAN PENJUALAN - TOKO SALFA'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Periode: {dari or 'Semua'} s/d {sampai or 'Semua'}"
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['No', 'Nama Produk', 'Jumlah Terjual', 'Total Penjualan (Rp)', 'Total Profit (Rp)']
    header_fill = PatternFill(start_color='1A5276', end_color='1A5276', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    row = 5
    for i, (nama, val) in enumerate(ringkasan.items(), 1):
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=nama)
        ws.cell(row=row, column=3, value=val['jumlah'])
        ws.cell(row=row, column=4, value=round(val['total'], 0))
        ws.cell(row=row, column=5, value=round(val['profit'], 0))
        if row % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
        row += 1

    ws.cell(row=row+1, column=3, value='Grand Total').font = Font(bold=True)
    ws.cell(row=row+1, column=4, value=float(grand_total)).font = Font(bold=True)
    ws.cell(row=row+2, column=3, value='Total Profit').font = Font(bold=True)
    ws.cell(row=row+2, column=5, value=float(total_profit)).font = Font(bold=True)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="laporan_penjualan.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pdf(request):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    import io

    transaksi, dari, sampai = get_transaksi_filter(request)
    ringkasan, grand_total, total_profit = get_ringkasan(transaksi)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=10, spaceAfter=12, textColor=colors.grey)
    elements.append(Paragraph('LAPORAN PENJUALAN - TOKO SALFA', title_style))
    elements.append(Paragraph(f"Periode: {dari or 'Semua'} s/d {sampai or 'Semua'}", sub_style))

    data = [['No', 'Nama Produk', 'Jumlah Terjual', 'Total Penjualan (Rp)', 'Total Profit (Rp)']]
    for i, (nama, val) in enumerate(ringkasan.items(), 1):
        data.append([
            str(i),
            nama,
            str(val['jumlah']),
            f"{val['total']:,.0f}",
            f"{val['profit']:,.0f}",
        ])

    data.append(['', '', '', '', ''])
    data.append(['', '', 'Grand Total', f"{float(grand_total):,.0f}".replace(',', '.'), ''])
    data.append(['', '', 'Total Profit', '', f"{float(total_profit):,.0f}".replace(',', '.')])

    table = Table(data, colWidths=[1*cm, 7*cm, 3.5*cm, 4*cm, 4*cm], repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1A5276')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('ALIGN', (0,0), (0,-1), 'CENTER'),
        ('ROWBACKGROUNDS', (0,1), (-1,-4), [colors.white, colors.HexColor('#EBF5FB')]),
        ('GRID', (0,0), (-1,-4), 0.5, colors.grey),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-2), (-1,-1), colors.HexColor('#D6EAF8')),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="laporan_penjualan.pdf"'
    return response